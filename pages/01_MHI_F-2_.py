import streamlit as st
import pandas as pd
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ファイル名を固定
file_name = "./data/MHI_F-2.xlsx"

# Excelをダウンロードする関数
def generate_excel_with_hyperlinks(df):
    # 新規Excelワークブックを作成
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # ヘッダーを設定
    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")  # ヘッダを左揃え

    # データを追加し、ハイパーリンクを設定
    for row_num, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left")  # 左揃え

            # 2列目（パス列）をハイパーリンクに変換
            if col_num == 2 and os.path.exists(value):
                cell.hyperlink = value
                cell.style = "Hyperlink"
                # B列セルの背景色を設定
                cell.fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    # B列とA列の幅を調整
    column_widths = []
    for row in dataframe_to_rows(df, index=False, header=True):
        for i, cell_value in enumerate(row):
            if len(column_widths) > i:
                column_widths[i] = max(column_widths[i], len(str(cell_value)))
            else:
                column_widths.append(len(str(cell_value)))

    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)  # 列をA, B, C...に変換
        if i == 1:
            ws.column_dimensions[column_letter].width = (width * 2) + 2  # A列の幅を調整
        elif i == 2:
            ws.column_dimensions[column_letter].width = (width * 1.5) + 2  # B列の幅を調整
        else:
            ws.column_dimensions[column_letter].width = width + 2  # 他の列の幅を調整

    # Excelデータをバイトストリームに保存
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data

# タイトル
item_name = "MHI F-2"
st.title(item_name)
#st.markdown("<h1 style='text-align: left;'>KHI P-1 C-2</h1>", unsafe_allow_html=True)

try:
    # Excelファイルを読み込む
    df = pd.read_excel(file_name)

    # データ表示
    st.write("\u8aad\u307f\u8fbc\u3093\u3060\u30c7ー\u30bf:")

    # CSSを追加して1列目のセルサイズを大きくし、内容が改行されないようにする
    st.markdown(
        """
        <style>
        table th {
            text-align: left !important;
        }
        table td {
            text-align: left !important;
            white-space: nowrap !important; /* 改行を防止 */
        }
        table td:first-child {
            width: 200px !important; /* 1列目のセルサイズを大きく設定 */
        }
        table td:nth-child(2) {
            max-width: 150px !important; /* 2列目の最大幅を制限 */
            overflow: hidden !important; /* 内容を切り捨て */
            text-overflow: ellipsis !important; /* 省略記号を追加 */
            white-space: nowrap !important; /* 改行を防止 */
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    # DataFrameをHTMLとして表示
    st.markdown(
        df.to_html(index=False, escape=False), 
        unsafe_allow_html=True
    )

    # ダウンロード用リンクを作成
    excel_data = generate_excel_with_hyperlinks(df)

    st.download_button(
        label="Excel\u5f62\u5f0f\u3067\u30c0\u30a6\u30f3\u30ed\u30fc\u30c9",
        data=excel_data,
        file_name="downloaded_" + item_name + ".xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except FileNotFoundError:
    st.error(f"指定されたファイル '{file_name}' が見つかりません。")
except Exception as e:
    st.error(f"エラーが発生しました: {str(e)}")





